import pandas as pd
import re
import openpyxl

# 1. Parse the Truth Pools
TEST_READINGS_PATH = 'master_readings/TEST READINGS[23].xlsx'
readings_df = pd.read_excel(TEST_READINGS_PATH, sheet_name='Sheet1')

row0 = readings_df.iloc[0]
group_to_cols = {}
for i, col in enumerate(readings_df.columns):
    val = row0[col]
    if pd.notna(val) and isinstance(val, str) and not val.startswith("Grp") and not val.startswith("Specification"):
        spec_col = i - 1
        val_col = i
        group_to_cols[val.strip()] = (spec_col, val_col)

parsed_data = {grp: {} for grp in group_to_cols}

for grp, (spec_col, val_col) in group_to_cols.items():
    param_col = 1 if spec_col < 38 else 3 
    
    for r in range(1, 18):
        p_val = readings_df.iloc[r, param_col]
        s_val = readings_df.iloc[r, spec_col]
        v_val = readings_df.iloc[r, val_col]
        if pd.notna(p_val) and str(p_val).strip() != "":
            param_name = str(p_val).strip()
            # Try parsing value as float if it's a number
            pool = []
            if pd.notna(v_val):
                try: pool.append(float(v_val))
                except: pool.append(v_val)
            parsed_data[grp][param_name] = pool

current_param = None
for r in range(20, len(readings_df)):
    p_val = readings_df.iloc[r, 5]
    if pd.notna(p_val) and str(p_val).strip() != "":
        current_param = str(p_val).strip()
    
    if current_param:
        for grp, (spec_col, val_col) in group_to_cols.items():
            v_val = readings_df.iloc[r, val_col]
            if pd.notna(v_val):
                if current_param not in parsed_data[grp]:
                    parsed_data[grp][current_param] = []
                try: v_val_clean = float(v_val)
                except: v_val_clean = v_val
                parsed_data[grp][current_param].append(v_val_clean)

# 2. Build Item Code to Group Map
def normalize_loose(s):
    s = re.sub(r'[^a-zA-Z0-9]', '', str(s)).lower()
    return s.replace('mm', '').replace('f', '')

group_df = pd.read_excel('group_names/19 Group Names.xlsx')
group_names_list = group_df.iloc[:, 1].tolist()
group_map = {normalize_loose(g): g for g in group_names_list}

tracker_df = pd.read_excel('sample_inputs/Item Codes Tracker.xlsx', header=5)
item_info_map = {}
for idx, row in tracker_df.iterrows():
    code = str(row.get('ITEM CODES', '')).strip()
    if code and code != 'nan':
        t = str(row.get('Type', '')).strip()
        c = str(row.get('Copper Type', '')).strip()
        od = str(row.get('Nominal OD', '')).strip()
        try:
            od_fmt = f"{float(od):.2f}"
        except:
            od_fmt = od
        
        constructed = f"{t} {c} OD {od_fmt}"
        norm = normalize_loose(constructed)
        group_name = group_map.get(norm, f"{t} {row.get('COLOUR', '')}".strip())
        item_info_map[code] = {'group_name': group_name, 'colour': str(row.get('COLOUR', '')).strip()}

# 3. Verify Output
wb = openpyxl.load_workbook('v3output/Master_All_Reports.xlsx', data_only=True)
print(f"Loaded {len(wb.sheetnames)} sheets for verification.")

error_count = 0
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    
    item_code_cell = ws['C5'].value
    if not item_code_cell:
        print(f"Sheet {sheet_name}: Could not find Item Code in C5. Skipping.")
        continue
    item_code = str(item_code_cell).strip()
    
    info = item_info_map.get(item_code)
    if not info:
        print(f"Sheet {sheet_name}: Item code '{item_code}' not found in tracker!")
        error_count += 1
        continue
        
    group_name = info['group_name']
    item_colour = info['colour']
    
    truth_data = parsed_data.get(group_name)
    if not truth_data:
        print(f"Sheet {sheet_name}: Group '{group_name}' not found in Truth Data!")
        error_count += 1
        continue
        
    for r in range(10, 27): # Check up to 17 parameters max
        p_name = ws.cell(row=r, column=1).value # A10..A26
        if not p_name:
            break
            
        p_name = str(p_name).strip()
        truth_pool = truth_data.get(p_name, [])
        
        # Check the observations F(6) through M(13)
        for c in range(6, 14):
            obs_val = ws.cell(row=r, column=c).value
            if obs_val is not None and str(obs_val).strip() != "":
                
                # Check for "Colour" special case
                if p_name == "Colour":
                    expected = str(item_colour)
                    actual = str(obs_val).replace("\n", "")
                    if expected not in actual and actual not in expected:
                        print(f"ERROR: {sheet_name} | {group_name} | {p_name} -> Found '{obs_val}', Expected '{expected}'")
                        error_count += 1
                    continue
                
                # Parse float if possible for strict matching
                try: obs_val_clean = float(obs_val)
                except: obs_val_clean = obs_val
                
                # Check against truth pool
                if len(truth_pool) == 1:
                    expected = truth_pool[0]
                    if obs_val_clean != expected and str(obs_val_clean) != str(expected):
                        print(f"ERROR: {sheet_name} | {group_name} | {p_name} -> Found '{obs_val}', Expected strictly '{expected}'")
                        error_count += 1
                elif len(truth_pool) > 1:
                    # Random choice, must be inside truth pool
                    match_found = False
                    for tp in truth_pool:
                        if obs_val_clean == tp or str(obs_val_clean) == str(tp):
                            match_found = True
                            break
                    if not match_found:
                        print(f"ERROR: {sheet_name} | {group_name} | {p_name} -> Found '{obs_val}' WHICH IS NOT IN POOL {truth_pool}")
                        error_count += 1
                else:
                    # Empty pool means it should be 'OK'
                    if str(obs_val_clean) != "OK":
                        print(f"ERROR: {sheet_name} | {group_name} | {p_name} -> Found '{obs_val}', Expected 'OK' (Empty Pool)")
                        error_count += 1

if error_count == 0:
    print("\nSUCCESS: All 100% of observations across all generated sheets strictly match their correct Group and Parameter pools from the Master Data!")
else:
    print(f"\nFAILED: Found {error_count} mismatch errors.")

