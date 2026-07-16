import os
import pandas as pd
import random
from datetime import datetime, timedelta
import math
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText

BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))
ITEM_CODES_TRACKER_PATH = os.path.join(BASE_DIR, "sample_inputs/Item Codes Tracker.xlsx")
TEST_READINGS_PATH = os.path.join(BASE_DIR, "master_readings/TEST READINGS[23].xlsx")
GROUP_NAMES_PATH = os.path.join(BASE_DIR, "group_names/19 Group Names.xlsx")

STAMP_PATH = os.path.join(BASE_DIR, "images", "stamp-removebg-preview.png")
ELENTEC_LOGO = os.path.join(BASE_DIR, "images", "WhatsApp Image 2026-06-25 at 00.15.42.jpeg")
VARDHAN_LOGO = os.path.join(BASE_DIR, "images", "WhatsApp Image 2026-06-25 at 00.16.24.jpeg")

def normalize_loose(s):
    s = re.sub(r'[^a-zA-Z0-9]', '', str(s)).lower()
    return s.replace('mm', '').replace('f', '')

category_map = {
    "No of strands": "A", "Strands dia": "C", "Insulation dia": "C", "Conductor resistance": "B",
    "Insulation thickness": "C", "Insulation thickness (Nom.)": "C", "Temperature rating": "D",
    "Temprature rating": "D", "Voltage rating": "D", "Volume Resistivity": "B", "Withstand voltage": "B",
    "Printing over the cable": "A", "Insulator Elongation": "B", "Conductor Elongation": "B",
    "Tensile Strength": "B", "Shrinkage": "B", "Spark Testing": "B", "Colour": "A", "Appearance": "A"
}

def build_worksheet(ws, data):
    ws.sheet_view.showGridLines = False

    # Print setup for perfect PDF export
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.print_options.horizontalCentered = True
    
    # Styles
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    calibri_9 = Font(name='Calibri', size=9)
    calibri_11 = Font(name='Calibri', size=11)
    bold_11 = Font(name='Calibri', size=11, bold=True)
    title_font = Font(name='Calibri', size=16, bold=True)
    
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center', wrap_text=True)
    top_left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
    top_center_align = Alignment(horizontal='center', vertical='top', wrap_text=True)
    
    gray_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    
    # Set default font for all cells to Calibri 11
    for row in range(1, 40):
        for col in range(1, 27):
            ws.cell(row=row, column=col).font = calibri_11

    # Widths
    widths = [3.8, 13.8, 6.8, 9.5, 8.3, 7.7, 7.7, 7.7, 7.7, 7.7, 7.7, 7.7, 7.7, 7.0, 11.5, 7.7, 7.7, 7.7, 7.7, 8.8, 7.7, 7.7, 7.7, 7.7, 8.0, 8.2]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64+i)].width = w

    # Heights
    heights = {
        1: 24.8, 2: 20.2, 3: 22.5, 4: 22.5, 5: 22.5, 6: 22.5, 7: 27.8, 8: 15.0, 9: 31.5,
        26: 73.5, 27: 18.0, 28: 23.2, 29: 15.0, 30: 15.0, 31: 15.0, 32: 15.0, 33: 15.0, 34: 16.0, 35: 15.8
    }
    for r in range(1, 36):
        if r in heights:
            ws.row_dimensions[r].height = heights[r]
        elif 10 <= r <= 25:
            ws.row_dimensions[r].height = 27.8
            
    # Helper to merge and set value/style to top-left
    def write_cell(range_str, value, font=None, alignment=None, fill=None):
        if ":" in range_str:
            ws.merge_cells(range_str)
            cells = ws[range_str]
            tl = range_str.split(':')[0]
            ws[tl].value = value
            for row in cells:
                for cell in row:
                    if font: cell.font = font
                    if alignment: cell.alignment = alignment
                    if fill: cell.fill = fill
        else:
            cell = ws[range_str]
            if value is not None:
                cell.value = value
            if font: cell.font = font
            if alignment: cell.alignment = alignment
            if fill: cell.fill = fill
        return cell

    def set_outer_border(ws, cell_range):
        rows = ws[cell_range]
        for r_idx, row in enumerate(rows):
            for c_idx, cell in enumerate(row):
                t = thin_border.top if r_idx == 0 else None
                b = thin_border.bottom if r_idx == len(rows) - 1 else None
                l = thin_border.left if c_idx == 0 else None
                r = thin_border.right if c_idx == len(row) - 1 else None
                cell.border = Border(top=t, bottom=b, left=l, right=r)

    # Row 1-2
    ws.merge_cells('A1:B2')
    set_outer_border(ws, 'A1:B2')
    ws.merge_cells('C1:D2')
    set_outer_border(ws, 'C1:D2')
    set_outer_border(ws, 'E1:Z2')
    if os.path.exists(ELENTEC_LOGO):
        try:
            img = Image(ELENTEC_LOGO)
            scale = min(100 / img.width, 55 / img.height)
            img.width = int(img.width * scale)
            img.height = int(img.height * scale)
            ws.add_image(img, 'A1')
        except: pass

    if os.path.exists(VARDHAN_LOGO):
        try:
            img2 = Image(VARDHAN_LOGO)
            scale = min(60 / img2.width, 55 / img2.height)
            img2.width = int(img2.width * scale)
            img2.height = int(img2.height * scale)
            ws.add_image(img2, 'C1')
        except: pass

    write_cell('E1:Z2', "Raw Material Quality Inspection Plan", title_font, center_align)

    # Row 3
    write_cell('A3:N3', "Supplier- VARDHAN CABLE INDUSTRIES", bold_11, center_align)
    write_cell('O3:Z3', "EIPL", bold_11, center_align)

    # Row 4
    write_cell('A4:B4', "Item Description", calibri_11, left_align)
    write_cell('C4:E4', data['item_description'], bold_11, center_align)
    write_cell('F4:H4', "Inspected By:", calibri_11, left_align)
    write_cell('I4:N4', "AJAY", bold_11, center_align)
    write_cell('O4', "EI Part No", calibri_11, left_align)
    write_cell('P4:Z4', "", calibri_11, left_align)

    # Row 5
    write_cell('A5:B5', "Item Code/Part No", calibri_11, left_align)
    write_cell('C5:E5', data['item_code'], bold_11, center_align)
    write_cell('F5:H5', "Approved BY:", calibri_11, left_align)
    write_cell('I5:N5', "SAURABH", bold_11, center_align)
    write_cell('O5', "GRN NO", calibri_11, left_align)
    write_cell('P5:S5', "", calibri_11, left_align)
    write_cell('T5', "GRN Date", calibri_11, left_align)
    write_cell('U5:Z5', "", calibri_11, left_align)

    # Row 6
    write_cell('A6:B6', "Lot Size", calibri_11, left_align)
    write_cell('C6:D6', data.get('lot_size_metres', "M M"), bold_11, center_align)
    write_cell('E6', f"{data['coils']} coils", bold_11, center_align)
    write_cell('F6:H6', "Insp Date", calibri_11, left_align)
    write_cell('I6:N6', data['insp_date'], bold_11, center_align)
    write_cell('O6', "Insp Date", calibri_11, left_align)
    write_cell('P6:S6', "", calibri_11, left_align)
    write_cell('T6', "Insp By", calibri_11, left_align)
    write_cell('U6:Z6', "", calibri_11, left_align)

    # Row 7
    write_cell('A7:B7', "Lot No", calibri_11, left_align)
    write_cell('C7:E7', data['lot_date'], bold_11, center_align)
    write_cell('F7:H7', "Invoice No ,Date", calibri_11, left_align)
    write_cell('I7:K7', data['invoice_no'], bold_11, center_align)
    write_cell('L7:N7', data['invoice_date'], bold_11, center_align)
    calibri_24 = Font(name='Arial', size=32)
    # Pass
    write_cell('O7:P7', "Pass", calibri_11, left_align)
    write_cell('Q7', "▭", calibri_24, center_align)

    # Pass Under deviation
    write_cell('R7:U7', "Pass Under deviation", calibri_11, center_align)
    write_cell('V7', "▭", calibri_24, center_align)

    # Rejected
    write_cell('W7:Y7', "Rejected", calibri_11, center_align)
    write_cell('Z7', "▭", calibri_24, center_align)

    # Row 8 & 9 Merged for first few columns
    write_cell('A8:B9', "Parameter", bold_11, center_align, gray_fill)
    write_cell('C8:D9', "Specification", bold_11, center_align, gray_fill)
    write_cell('E8:E9', "Sample\nSize", bold_11, center_align, gray_fill)
    write_cell('F8:M8', "Observation", bold_11, center_align, gray_fill)
    write_cell('N8', "", calibri_11, center_align, gray_fill)
    write_cell('O8:Z8', "Observation", bold_11, center_align, gray_fill)

    # Row 9 remainder
    for i, col in enumerate("FGHIJKLM"):
        write_cell(f"{col}9", i+1, bold_11, center_align, gray_fill)
    write_cell('N9', "Status", bold_11, center_align, gray_fill)
    write_cell('O9', "Inspection Category", bold_11, center_align, gray_fill)
    write_cell('P9', "Sample Size", bold_11, center_align, gray_fill)
    for i, col in enumerate("QRSTUVWX"):
        write_cell(f"{col}9", i+1, bold_11, center_align, gray_fill)
    write_cell('Y9', "Method", bold_11, center_align, gray_fill)
    write_cell('Z9', "Status", bold_11, center_align, gray_fill)

    # Data rows 10 to 26
    current_row = 10
    for row_data in data['readings']:
        p_name = row_data['parameter']
        write_cell(f"A{current_row}:B{current_row}", p_name, bold_11, center_align)
        write_cell(f"C{current_row}:D{current_row}", row_data['specification'], calibri_11, center_align)
        
        sample_size = "8"
        if p_name == "Spark Testing":
            sample_size = "100%"
        elif p_name == "Appearance":
            sample_size = "8"
        write_cell(f"E{current_row}", sample_size, calibri_11, center_align)
        
        # Observations F-M
        if p_name == "Appearance":
            ws.row_dimensions[current_row].height = 73.5
            for i in range(data['coils']):
                if i < 8:
                    write_cell(f"{chr(70+i)}{current_row}", "OK", calibri_11, center_align)
        elif p_name == "Colour":
            ws.row_dimensions[current_row].height = 45.0
            for i, obs_val in enumerate(row_data['obs']):
                if i < data['coils']:
                    formatted_obs = str(obs_val).replace('/', '/\n')
                    write_cell(f"{chr(70+i)}{current_row}", formatted_obs, calibri_9, center_align)
        else:
            for i, obs_val in enumerate(row_data['obs']):
                if i < data['coils']:
                    write_cell(f"{chr(70+i)}{current_row}", str(obs_val), calibri_11, center_align)
        
        write_cell(f"N{current_row}", "OK", bold_11, center_align)
        write_cell(f"O{current_row}", row_data['insp_category'], calibri_11, center_align)
        
        if p_name in ["Colour", "Appearance"]:
            write_cell(f"Y{current_row}", "Visual", calibri_11, center_align)
            
        current_row += 1

    # Row 27
    write_cell('A27:N27', "", calibri_11, center_align)
    write_cell('S27:Z27', "Inspection category:", bold_11, left_align)

    # Row 28
    write_cell('A28:G30', "The product must be RoHS/ Phthalates Free compliant and have a vaild Certificate.", bold_11, center_align)
    write_cell('H28:L30', "Refer ISO-2859 Inspection Standards", bold_11, center_align)
    write_cell('M28', "Yes", calibri_11, center_align)
    write_cell('N28', "NO", calibri_11, center_align)
    write_cell('O28:P28', "OK", calibri_11, center_align)
    write_cell('Q28:R28', "Rej", calibri_11, center_align)
    write_cell('S28:Z28', "A     Appearance Visual", calibri_11, left_align)

    # Row 29
    tick_font = Font(name='Calibri', size=24)
    write_cell('M29:M30', "✓", tick_font, center_align)
    write_cell('N29:N30', "", calibri_11, center_align)
    write_cell('O29:P30', "", calibri_11, center_align)
    write_cell('Q29:R30', "", calibri_11, center_align)
    write_cell('S29:Z29', "B     Electrical (High voltage tester)", calibri_11, left_align)

    # Row 30
    write_cell('S30:Z30', "C     Dimension(Mich,vernier,quick mini)", calibri_11, left_align)

    # Row 31
    write_cell('A31:R34', "Remarks:\nHighlighted in Bold Letters are the CTQs", calibri_11, top_left_align)
    write_cell('S31:Z31', "D     Supplier readings", calibri_11, left_align)

    # Row 32
    write_cell('S32:Z32', "E      Fitment", calibri_11, left_align)

    # Row 33
    write_cell('S33:Z33', "sampling plan -ISO 2859, AQL-1.5G1", calibri_11, left_align)

    # Row 34
    write_cell('S34:Z34', "Note :8 readings will be recorded", calibri_11, left_align)

    # Row 35
    write_cell('W35:Z35', "EF 119/3/17.03.2018", calibri_11, right_align)

    if os.path.exists(STAMP_PATH):
        try:
            img3 = Image(STAMP_PATH)
            scale = min(500 / img3.width, 180 / img3.height)
            img3.width = int(img3.width * scale)
            img3.height = int(img3.height * scale)
            ws.add_image(img3, 'H29')
        except: pass

    # Apply Borders from Row 3 to 35, handling merged cells properly to prevent corruption
    merged_cells_coords = set()
    for merged_range in ws.merged_cells.ranges:
        for row in ws[merged_range.coord]:
            for cell in row:
                merged_cells_coords.add(cell.coordinate)
                
    for r in range(3, 36):
        if r == 35: continue
        for c in range(1, 27):
            cell = ws.cell(row=r, column=c)
            if cell.coordinate not in merged_cells_coords:
                cell.border = thin_border
                
    for merged_range in ws.merged_cells.ranges:
        bounds = merged_range.bounds
        if bounds[1] >= 3 and bounds[1] <= 34:
            rows = list(ws[merged_range.coord])
            top_border = thin_border.top
            bottom_border = thin_border.bottom
            left_border = thin_border.left
            right_border = thin_border.right
            for cell in rows[0]:
                cell.border = Border(top=top_border, bottom=cell.border.bottom, left=cell.border.left, right=cell.border.right)
            for cell in rows[-1]:
                cell.border = Border(bottom=bottom_border, top=cell.border.top, left=cell.border.left, right=cell.border.right)
            for row in rows:
                row[0].border = Border(left=left_border, right=row[0].border.right, top=row[0].border.top, bottom=row[0].border.bottom)
                row[-1].border = Border(right=right_border, left=row[-1].border.left, top=row[-1].border.top, bottom=row[-1].border.bottom)
            
    # Re-apply custom borders for Row 7 checkboxes
    ws['O7'].border = Border(left=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style=None))
    ws['P7'].border = Border(left=Side(style=None), top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style=None))
    ws['Q7'].border = Border(left=Side(style=None), top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'))

    ws['R7'].border = Border(left=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style=None))
    for col in ['S', 'T', 'U']:
        ws[f'{col}7'].border = Border(left=Side(style=None), top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style=None))
    ws['V7'].border = Border(left=Side(style=None), top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'))

    ws['W7'].border = Border(left=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style=None))
    for col in ['X', 'Y']:
        ws[f'{col}7'].border = Border(left=Side(style=None), top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style=None))
    ws['Z7'].border = Border(left=Side(style=None), top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'))





def process_packing_slip(input_path, output_dir, output_mode="single"):
    os.makedirs(output_dir, exist_ok=True)
    
    # 1. Load mappings
    try:
        group_df = pd.read_excel(GROUP_NAMES_PATH)
        group_names_list = group_df.iloc[:, 1].tolist()
    except Exception: group_names_list = []
    
    group_map = {normalize_loose(g): g for g in group_names_list}
    
    tracker_df = pd.read_excel(ITEM_CODES_TRACKER_PATH, header=5)
    item_info_map = {}
    for idx, row in tracker_df.iterrows():
        code = str(row.get('ITEM CODES', '')).strip()
        if code and code != 'nan':
            t = str(row.get('Type', '')).strip()
            c = str(row.get('Copper Type', '')).strip()
            od = str(row.get('Nominal OD', '')).strip()
            try: od_fmt = f"{float(od):.2f}"
            except: od_fmt = od
            constructed = f"{t} {c} OD {od_fmt}"
            norm = normalize_loose(constructed)
            group_name = group_map.get(norm, f"{t} {row.get('COLOUR', '')}".strip())
            item_info_map[code] = {'type': t, 'colour': str(row.get('COLOUR', '')).strip(), 'copper_type': c, 'group_name': group_name}
            
    # 2. Load pools
    readings_df = pd.read_excel(TEST_READINGS_PATH, sheet_name='Sheet1')
    row0 = readings_df.iloc[0]
    group_to_cols = {}
    for i, col in enumerate(readings_df.columns):
        val = row0[col]
        if pd.notna(val) and isinstance(val, str) and not val.startswith("Grp") and not val.startswith("Specification"):
            group_to_cols[val.strip()] = (i - 1, i)
            
    parsed_data = {grp: {} for grp in group_to_cols}
    for grp, (spec_col, val_col) in group_to_cols.items():
        param_col = 1 if spec_col < 38 else 3 
        for r in range(1, 18):
            p_val = readings_df.iloc[r, param_col]
            s_val = readings_df.iloc[r, spec_col]
            v_val = readings_df.iloc[r, val_col]
            if pd.notna(p_val) and str(p_val).strip() != "":
                pool = [v_val] if pd.notna(v_val) else []
                parsed_data[grp][str(p_val).strip()] = {'spec': str(s_val).strip() if pd.notna(s_val) else "", 'pool': pool}
                
    current_param = None
    for r in range(20, len(readings_df)):
        p_val = readings_df.iloc[r, 5]
        if pd.notna(p_val) and str(p_val).strip() != "":
            current_param = str(p_val).strip()
        if current_param:
            for grp, (spec_col, val_col) in group_to_cols.items():
                v_val = readings_df.iloc[r, val_col]
                if pd.notna(v_val):
                    if current_param not in parsed_data[grp]: parsed_data[grp][current_param] = {'spec': '', 'pool': []}
                    parsed_data[grp][current_param]['pool'].append(v_val)
                    
    # 3. Process Packing Slip Items
    ps_df = pd.read_excel(input_path)
    current_invoice_no = None
    current_invoice_date = None
    
    # First, collect all items and sort by Group if in group mode
    items_to_process = []
    
    for idx, row in ps_df.iterrows():
        c2 = str(row.get('Unnamed: 2', '')).strip()
        c3 = str(row.get('Unnamed: 3', '')).strip()
        c4 = str(row.get('Unnamed: 4', '')).strip()
        if "INVOICE NO" in c2:
            current_invoice_no = c3
            if "Dt:" in c4: current_invoice_date = c4.replace("Dt:", "").strip()
                
        c1 = str(row.get('Unnamed: 1', '')).strip()
        if c1.isdigit() and current_invoice_no:
            item_code = c2
            coils_val = row.get('Unnamed: 7', float('nan'))
            if pd.isna(coils_val): continue
            num_coils = int(math.ceil(float(coils_val)))
            length_val = row.get('Unnamed: 8', float('nan'))
            lot_size_metres = f"{int(float(length_val))} m" if pd.notna(length_val) else "M M"
            
            info = item_info_map.get(item_code, {})
            full_desc = info.get('group_name', 'Unknown Group')
            
            items_to_process.append({
                'item_code': item_code,
                'num_coils': num_coils,
                'lot_size_metres': lot_size_metres,
                'invoice_no': current_invoice_no,
                'invoice_date': current_invoice_date,
                'copper': info.get('copper_type', ''),
                'colour': info.get('colour', ''),
                'group': full_desc
            })
            
    if output_mode == "group":
        items_to_process.sort(key=lambda x: x['group'])
        
    master_wb = Workbook()
    master_ws_idx = 0
    group_mapping = {}
    current_group = None
    group_start_page = 0
    
    for item in items_to_process:
        # Track Group boundaries
        if item['group'] != current_group:
            if current_group is not None:
                group_mapping[current_group] = (group_start_page, master_ws_idx - 1)
            current_group = item['group']
            group_start_page = master_ws_idx
            
            if output_mode == "group":
                safe_group = current_group.replace("/", "_").replace(":", "_").strip()
                os.makedirs(os.path.join(output_dir, safe_group), exist_ok=True)
                
        try: inv_dt = datetime.strptime(item['invoice_date'], "%d-%m-%Y")
        except: inv_dt = datetime.now()
        insp_dt = inv_dt - timedelta(days=1)
        lot_dt = inv_dt - timedelta(days=2)
        
        cols_to_fill = min(8, item['num_coils'])
        
        readings_table = []
        group_data = parsed_data.get(item['group'], {})
        for p_name, p_data in group_data.items():
            spec = p_data.get('spec', '')
            pool = p_data.get('pool', [])
            cat = category_map.get(p_name, "A")
            obs = []
            
            # Special case overrides for parameters missing from the master sheet
            if p_name == "Printing over the cable":
                if not spec: spec = "NA"
                obs = ['NA'] * 8
            elif p_name == "Colour": 
                obs = [item['colour']] * 8
            elif len(pool) == 1: 
                obs = [pool[0]] * 8
            elif len(pool) > 1: 
                obs = [random.choice(pool) for _ in range(8)]
            else: 
                obs = ['OK'] * 8
                
            for i in range(cols_to_fill, 8): obs[i] = ""
            sample_size = "100%" if "Spark" in p_name else "8"
            readings_table.append({'parameter': p_name, 'specification': spec, 'sample_size': sample_size, 'obs': obs, 'status1': "OK" if cols_to_fill > 0 else "", 'insp_category': cat})
            
        data = {
            'item_description': item['group'],
            'item_code': item['item_code'],
            'coils': item['num_coils'],
            'lot_size_metres': item['lot_size_metres'],
            'insp_date': insp_dt.strftime("%d/%m/%Y"),
            'lot_date': lot_dt.strftime("%d/%m/%Y"),
            'invoice_no': item['invoice_no'],
            'invoice_date': item['invoice_date'].replace("-", "/"),
            'readings': readings_table
        }
        
        # Individual Workbook
        safe_item_code = item['item_code'].replace("/", "_").replace("\\\\", "_")
        if output_mode == "group":
            safe_group = item['group'].replace("/", "_").replace(":", "_").strip()
            excel_path = os.path.join(output_dir, safe_group, f"Inv_{item['invoice_no']}_{safe_item_code}.xlsx")
        else:
            excel_path = os.path.join(output_dir, f"Inv_{item['invoice_no']}_{safe_item_code}.xlsx")
            
        ind_wb = Workbook()
        build_worksheet(ind_wb.active, data)
        ind_wb.save(excel_path)
        
        # Master Workbook
        m_ws = master_wb.active if master_ws_idx == 0 else master_wb.create_sheet(title=f"{safe_item_code[:30]}_{master_ws_idx}")
        build_worksheet(m_ws, data)
        master_ws_idx += 1
        
    if current_group is not None:
        group_mapping[current_group] = (group_start_page, master_ws_idx - 1)
        
    master_excel_path = os.path.join(output_dir, "Master_All_Reports.xlsx")
    master_wb.save(master_excel_path)
    
    return master_excel_path, master_ws_idx, group_mapping

