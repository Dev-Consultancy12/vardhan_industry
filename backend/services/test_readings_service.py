import os
import pandas as pd
import openpyxl

BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))
TEST_READINGS_PATH = os.path.join(BASE_DIR, "master_readings/TEST READINGS[23].xlsx")

def get_test_readings():
    """
    Parses the master Excel file and returns the structured data along with exact cell coordinates.
    Structure:
    {
        "Group Name": {
            "Parameter Name": {
                "spec": "Specification string",
                "pool": [
                    {"val": 1.75, "row": 28, "col": 7}, # openpyxl is 1-indexed
                    {"val": 1.76, "row": 29, "col": 7}
                ]
            }
        }
    }
    """
    try:
        df = pd.read_excel(TEST_READINGS_PATH, sheet_name='Sheet1')
    except Exception as e:
        raise Exception(f"Failed to read test readings: {e}")

    row0 = df.iloc[0]
    group_to_cols = {}
    # Find columns for each group
    for i, col in enumerate(df.columns):
        val = row0[col]
        if pd.notna(val) and isinstance(val, str) and not val.startswith("Grp") and not val.startswith("Specification"):
            spec_col = i - 1
            val_col = i
            group_to_cols[val.strip()] = (spec_col, val_col)

    parsed_data = {}
    for grp in group_to_cols:
        parsed_data[grp] = {}

    # Upper Section (Fixed Values)
    for grp, (spec_col, val_col) in group_to_cols.items():
        param_col = 1 if spec_col < 38 else 3 
        
        for r in range(1, 18):
            p_val = df.iloc[r, param_col]
            s_val = df.iloc[r, spec_col]
            v_val = df.iloc[r, val_col]
            
            if pd.notna(p_val) and str(p_val).strip() != "":
                param_name = str(p_val).strip()
                spec_str = str(s_val).strip() if pd.notna(s_val) else ""
                
                pool = []
                if pd.notna(v_val):
                    try: v_clean = float(v_val)
                    except: v_clean = v_val
                    # openpyxl row = r + 2 (because pandas r=0 is Excel row 2 when header is row 1)
                    pool.append({"val": v_clean, "row": r + 2, "col": val_col + 1})
                
                parsed_data[grp][param_name] = {'spec': spec_str, 'pool': pool}

    # Lower Section (Randomized Pools)
    current_param = None
    for r in range(20, len(df)):
        p_val = df.iloc[r, 5]
        if pd.notna(p_val) and str(p_val).strip() != "":
            current_param = str(p_val).strip()
        
        if current_param:
            for grp, (spec_col, val_col) in group_to_cols.items():
                v_val = df.iloc[r, val_col]
                if pd.notna(v_val):
                    if current_param not in parsed_data[grp]:
                        parsed_data[grp][current_param] = {'spec': '', 'pool': []}
                    
                    try: v_clean = float(v_val)
                    except: v_clean = v_val
                    
                    parsed_data[grp][current_param]['pool'].append({
                        "val": v_clean,
                        "row": r + 2,
                        "col": val_col + 1
                    })

    return parsed_data

def update_test_reading(updates):
    """
    Updates the Excel file with the provided values.
    `updates` is a list of dicts: [{"row": 28, "col": 7, "val": 1.78}]
    """
    try:
        wb = openpyxl.load_workbook(TEST_READINGS_PATH)
        ws = wb['Sheet1']
        
        for update in updates:
            r = update['row']
            c = update['col']
            v = update['val']
            # Cast back to string if it's not a number, but keep numbers as numbers for Excel
            try:
                numeric_v = float(v)
                # Check if it's actually an integer
                if numeric_v.is_integer():
                    ws.cell(row=r, column=c).value = int(numeric_v)
                else:
                    ws.cell(row=r, column=c).value = numeric_v
            except ValueError:
                ws.cell(row=r, column=c).value = str(v)
                
        wb.save(TEST_READINGS_PATH)
        return True
    except Exception as e:
        raise Exception(f"Failed to save Excel file: {e}")
