def add_borders(ws, thin_border):
    from openpyxl.styles import Border
    
    # First, get a set of all coordinates inside merged cells
    merged_cells_coords = set()
    for merged_range in ws.merged_cells.ranges:
        for row in ws[merged_range.coord]:
            for cell in row:
                merged_cells_coords.add(cell.coordinate)
                
    # Apply border to all non-merged cells
    for r in range(3, 36):
        if r == 35: continue
        for c in range(1, 27):
            cell = ws.cell(row=r, column=c)
            if cell.coordinate not in merged_cells_coords:
                cell.border = thin_border
                
    # Apply proper borders to merged ranges
    for merged_range in ws.merged_cells.ranges:
        # Check if the merged range is within our target rows (3 to 34)
        bounds = merged_range.bounds  # (min_col, min_row, max_col, max_row)
        if bounds[1] >= 3 and bounds[1] <= 34:
            rows = list(ws[merged_range.coord])
            
            top_border = thin_border.top
            bottom_border = thin_border.bottom
            left_border = thin_border.left
            right_border = thin_border.right
            
            for cell in rows[0]:
                cell.border = Border(top=top_border, bottom=cell.border.bottom, left=cell.border.left, right=cell.border.right)
            for cell in rows[-1]:
                cell.border = Border(bottom=bottom_border, top=cell.border.top, left=cell.border.left, right=cell.border.right)
            for row in rows:
                row[0].border = Border(left=left_border, right=row[0].border.right, top=row[0].border.top, bottom=row[0].border.bottom)
                row[-1].border = Border(right=right_border, left=row[-1].border.left, top=row[-1].border.top, bottom=row[-1].border.bottom)
