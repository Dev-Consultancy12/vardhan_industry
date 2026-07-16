from openpyxl import Workbook
from openpyxl.styles import Border, Side
wb = Workbook()
ws = wb.active
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
ws.merge_cells('A1:C3')
ws['A1'].border = thin_border
wb.save('test_merge.xlsx')
