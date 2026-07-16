import openpyxl
from openpyxl.styles import Border, Side

wb = openpyxl.Workbook()
ws = wb.active

thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

ws.merge_cells('A1:C3')

for r in range(1, 5):
    for c in range(1, 5):
        ws.cell(row=r, column=c).border = thin_border

wb.save('test_corrupt.xlsx')
