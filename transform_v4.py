import re

with open('backend/services/excel_generator_v4.py', 'r') as f:
    content = f.read()

# Remove the hardcoded PACKING_SLIP_PATH config
content = content.replace('PACKING_SLIP_PATH = "sample_inputs/Packing Slip.xlsx"', '')
content = content.replace('OUTPUT_DIR = os.path.join(BASE_DIR, "v3output")', '')

# Wrap the parsing and generation logic into process_packing_slip
top_imports = content[:content.find('# --- DATA EXTRACTION ---')]

data_extraction_and_logic = content[content.find('# --- DATA EXTRACTION ---'):]

# In data_extraction, we need to wrap everything from line 26 down into the function, EXCEPT build_worksheet which can stay outside
# Let's just create a new file structure.

build_ws_start = data_extraction_and_logic.find('def build_worksheet(ws, data):')
build_ws_end = data_extraction_and_logic.find('current_invoice_no = None')

build_worksheet_func = data_extraction_and_logic[build_ws_start:build_ws_end]

# We need to construct the new file
new_content = """import os
import pandas as pd
import random
from datetime import datetime, timedelta
import math
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText

BASE_DIR = "/Users/namanbansal/vardhan_industry_automation"
ITEM_CODES_TRACKER_PATH = os.path.join(BASE_DIR, "sample_inputs/Item Codes Tracker.xlsx")
TEST_READINGS_PATH = os.path.join(BASE_DIR, "master_readings/TEST READINGS[23].xlsx")
GROUP_NAMES_PATH = os.path.join(BASE_DIR, "group_names/19 Group Names.xlsx")

STAMP_PATH = os.path.join(BASE_DIR, "images", "stamp-removebg-preview.png")
ELENTEC_LOGO = os.path.join(BASE_DIR, "images", "WhatsApp Image 2026-06-25 at 00.15.42.jpeg")
VARDHAN_LOGO = os.path.join(BASE_DIR, "images", "WhatsApp Image 2026-06-25 at 00.16.24.jpeg")

def normalize_loose(s):
    s = re.sub(r'[^a-zA-Z0-9]', '', str(s)).lower()
    return s.replace('mm', '').replace('f', '')

category_map = {
    "No of strands": "A", "Strands dia": "C", "Insulation dia": "C", "Conductor resistance": "B",
    "Insulation thickness": "C", "Insulation thickness (Nom.)": "C", "Temperature rating": "D",
    "Temprature rating": "D", "Voltage rating": "D", "Volume Resistivity": "B", "Withstand voltage": "B",
    "Printing over the cable": "A", "Insulator Elongation": "B", "Conductor Elongation": "B",
    "Tensile Strength": "B", "Shrinkage": "B", "Spark Testing": "B", "Colour": "A", "Appearance": "A"
}

""" + build_worksheet_func + """

def process_packing_slip(input_path, output_dir, output_mode="single"):
    os.makedirs(output_dir, exist_ok=True)
    
    # 1. Load mappings
    try:
        group_df = pd.read_excel(GROUP_NAMES_PATH)
        group_names_list = group_df.iloc[:, 1].tolist()
    except Exception: group_names_list = []
    
    group_map = {normalize_loose(g): g for g in group_names_list}
    
    tracker_df = pd.read_excel(ITEM_CODES_TRACKER_PATH, header=5)
    item_info_map = {}
    for idx, row in tracker_df.iterrows():
        code = str(row.get('ITEM CODES', '')).strip()
        if code and code != 'nan':
            t = str(row.get('Type', '')).strip()
            c = str(row.get('Copper Type', '')).strip()
            od = str(row.get('Nominal OD', '')).strip()
            try: od_fmt = f"{float(od):.2f}"
            except: od_fmt = od
            constructed = f"{t} {c} OD {od_fmt}"
            norm = normalize_loose(constructed)
            group_name = group_map.get(norm, f"{t} {row.get('COLOUR', '')}".strip())
            item_info_map[code] = {'type': t, 'colour': str(row.get('COLOUR', '')).strip(), 'copper_type': c, 'group_name': group_name}
            
    # 2. Load pools
    readings_df = pd.read_excel(TEST_READINGS_PATH, sheet_name='Sheet1')
    row0 = readings_df.iloc[0]
    group_to_cols = {}
    for i, col in enumerate(readings_df.columns):
        val = row0[col]
        if pd.notna(val) and isinstance(val, str) and not val.startswith("Grp") and not val.startswith("Specification"):
            group_to_cols[val.strip()] = (i - 1, i)
            
    parsed_data = {grp: {} for grp in group_to_cols}
    for grp, (spec_col, val_col) in group_to_cols.items():
        param_col = 1 if spec_col < 38 else 3 
        for r in range(1, 18):
            p_val = readings_df.iloc[r, param_col]
            s_val = readings_df.iloc[r, spec_col]
            v_val = readings_df.iloc[r, val_col]
            if pd.notna(p_val) and str(p_val).strip() != "":
                pool = [v_val] if pd.notna(v_val) else []
                parsed_data[grp][str(p_val).strip()] = {'spec': str(s_val).strip() if pd.notna(s_val) else "", 'pool': pool}
                
    current_param = None
    for r in range(20, len(readings_df)):
        p_val = readings_df.iloc[r, 5]
        if pd.notna(p_val) and str(p_val).strip() != "":
            current_param = str(p_val).strip()
        if current_param:
            for grp, (spec_col, val_col) in group_to_cols.items():
                v_val = readings_df.iloc[r, val_col]
                if pd.notna(v_val):
                    if current_param not in parsed_data[grp]: parsed_data[grp][current_param] = {'spec': '', 'pool': []}
                    parsed_data[grp][current_param]['pool'].append(v_val)
                    
    # 3. Process Packing Slip Items
    ps_df = pd.read_excel(input_path)
    current_invoice_no = None
    current_invoice_date = None
    
    # First, collect all items and sort by Group if in group mode
    items_to_process = []
    
    for idx, row in ps_df.iterrows():
        c2 = str(row.get('Unnamed: 2', '')).strip()
        c3 = str(row.get('Unnamed: 3', '')).strip()
        c4 = str(row.get('Unnamed: 4', '')).strip()
        if "INVOICE NO" in c2:
            current_invoice_no = c3
            if "Dt:" in c4: current_invoice_date = c4.replace("Dt:", "").strip()
                
        c1 = str(row.get('Unnamed: 1', '')).strip()
        if c1.isdigit() and current_invoice_no:
            item_code = c2
            coils_val = row.get('Unnamed: 7', float('nan'))
            if pd.isna(coils_val): continue
            num_coils = int(math.ceil(float(coils_val)))
            length_val = row.get('Unnamed: 8', float('nan'))
            lot_size_metres = f"{int(float(length_val))} m" if pd.notna(length_val) else "M M"
            
            info = item_info_map.get(item_code, {})
            full_desc = info.get('group_name', 'Unknown Group')
            
            items_to_process.append({
                'item_code': item_code,
                'num_coils': num_coils,
                'lot_size_metres': lot_size_metres,
                'invoice_no': current_invoice_no,
                'invoice_date': current_invoice_date,
                'copper': info.get('copper_type', ''),
                'colour': info.get('colour', ''),
                'group': full_desc
            })
            
    if output_mode == "group":
        items_to_process.sort(key=lambda x: x['group'])
        
    master_wb = Workbook()
    master_ws_idx = 0
    group_mapping = {}
    current_group = None
    group_start_page = 0
    
    for item in items_to_process:
        # Track Group boundaries
        if item['group'] != current_group:
            if current_group is not None:
                group_mapping[current_group] = (group_start_page, master_ws_idx - 1)
            current_group = item['group']
            group_start_page = master_ws_idx
            
            if output_mode == "group":
                safe_group = current_group.replace("/", "_").replace(":", "_").strip()
                os.makedirs(os.path.join(output_dir, safe_group), exist_ok=True)
                
        try: inv_dt = datetime.strptime(item['invoice_date'], "%d-%m-%Y")
        except: inv_dt = datetime.now()
        insp_dt = inv_dt - timedelta(days=1)
        lot_dt = inv_dt - timedelta(days=2)
        
        cols_to_fill = min(8, item['num_coils'])
        
        readings_table = []
        group_data = parsed_data.get(item['group'], {})
        for p_name, p_data in group_data.items():
            spec = p_data.get('spec', '')
            pool = p_data.get('pool', [])
            cat = category_map.get(p_name, "A")
            obs = []
            if p_name == "Colour": obs = [item['colour']] * 8
            elif len(pool) == 1: obs = [pool[0]] * 8
            elif len(pool) > 1: obs = [random.choice(pool) for _ in range(8)]
            else: obs = ['OK'] * 8
                
            for i in range(cols_to_fill, 8): obs[i] = ""
            sample_size = "100%" if "Spark" in p_name else "8"
            readings_table.append({'parameter': p_name, 'specification': spec, 'sample_size': sample_size, 'obs': obs, 'status1': "OK" if cols_to_fill > 0 else "", 'insp_category': cat})
            
        data = {
            'item_description': item['group'],
            'item_code': item['item_code'],
            'coils': item['num_coils'],
            'lot_size_metres': item['lot_size_metres'],
            'insp_date': insp_dt.strftime("%d/%m/%Y"),
            'lot_date': lot_dt.strftime("%d/%m/%Y"),
            'invoice_no': item['invoice_no'],
            'invoice_date': item['invoice_date'].replace("-", "/"),
            'readings': readings_table
        }
        
        # Individual Workbook
        safe_item_code = item['item_code'].replace("/", "_").replace("\\", "_")
        if output_mode == "group":
            safe_group = item['group'].replace("/", "_").replace(":", "_").strip()
            excel_path = os.path.join(output_dir, safe_group, f"Inv_{item['invoice_no']}_{safe_item_code}.xlsx")
        else:
            excel_path = os.path.join(output_dir, f"Inv_{item['invoice_no']}_{safe_item_code}.xlsx")
            
        ind_wb = Workbook()
        build_worksheet(ind_wb.active, data)
        ind_wb.save(excel_path)
        
        # Master Workbook
        m_ws = master_wb.active if master_ws_idx == 0 else master_wb.create_sheet(title=f"{safe_item_code[:30]}_{master_ws_idx}")
        build_worksheet(m_ws, data)
        master_ws_idx += 1
        
    if current_group is not None:
        group_mapping[current_group] = (group_start_page, master_ws_idx - 1)
        
    master_excel_path = os.path.join(output_dir, "Master_All_Reports.xlsx")
    master_wb.save(master_excel_path)
    
    return master_excel_path, master_ws_idx, group_mapping

"""

with open('backend/services/excel_generator_v4.py', 'w') as f:
    f.write(new_content)

